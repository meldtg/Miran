from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import UserProfile, AccessCode, DayAssignment, Area, SupportProposal
from datetime import date, datetime
from django.utils.dateparse import parse_time, parse_date
from django.utils import timezone
from django.db import transaction
from django.conf import settings
from .services.telegram_logger import send_user_verified_log
from .services.telegram_notifier import send_excel_report
from django.db.models.functions import ExtractYear, ExtractMonth
from django.db.models import Count, Max
import os
from urllib.parse import urlparse


@api_view(["GET"])
def ping(_request):
    return Response({"status": "ok"})


@api_view(["POST"])
def init(request):
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)

    user_id = int(user_data["id"])
    profile, created = UserProfile.objects.get_or_create(
        user_id=user_id,
        defaults={
            "last_name": "",
            "first_name": "",
            "middle_name": "",
        },
    )
    # If no role assigned, deny access
    if not profile.role:
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    return Response({
        "ok": True,
        "created": created,
        "user": {
            "user_id": profile.user_id,
            "last_name": profile.last_name,
            "first_name": profile.first_name,
            "middle_name": profile.middle_name,
            "full_name": profile.full_name,
            "role": profile.role,
        }
    })

# Helper to build absolute public media URLs on fixed host
PUBLIC_MEDIA_HOST = os.getenv("PUBLIC_MEDIA_HOST", "https://miran-production.up.railway.app").rstrip("/")

def make_public_url(url: str) -> str:
    if not url:
        return url
    if url.startswith("http://") or url.startswith("https://"):
        p = urlparse(url)
        path_q = p.path + (("?" + p.query) if p.query else "")
        return f"{PUBLIC_MEDIA_HOST}{path_q}"
    # url is a path like /media/...
    return f"{PUBLIC_MEDIA_HOST}{url}"

@api_view(["POST"])
def check_code(request):
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)

    body = request.data or {}
    code = (body.get("code") or "").strip()
    if not code:
        return Response({"error": "code is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        ac = AccessCode.objects.get(code=code)
    except AccessCode.DoesNotExist:
        return Response({"error": "invalid code"}, status=status.HTTP_404_NOT_FOUND)

    user_id = int(user_data["id"])
    profile, _ = UserProfile.objects.get_or_create(
        user_id=user_id,
        defaults={"last_name": "", "first_name": "", "middle_name": ""},
    )
    profile.last_name = ac.last_name
    profile.first_name = ac.first_name
    profile.middle_name = ac.middle_name
    profile.role = ac.role
    profile.save(update_fields=["last_name", "first_name", "middle_name", "role", "updated_at"])

    # Send log to Telegram topic
    try:
        send_user_verified_log(user_data=user_data, full_name=profile.full_name)
    except Exception:
        pass

    return Response({
        "ok": True,
        "user": {
            "user_id": profile.user_id,
            "last_name": profile.last_name,
            "first_name": profile.first_name,
            "middle_name": profile.middle_name,
            "full_name": profile.full_name,
            "role": profile.role,
        }
    })


@api_view(["GET"])
def my_checker_days(request):
    """
    Возвращает дни (и участки) для текущего пользователя, на которые он назначен ответственным.
    Предназначено для роли 'checker', но вернёт пустой список для других ролей.
    Формат:
    {
      "ok": true,
      "days": [
        { "date": "2025-11-12", "areas": [ { "id": 1, "name": "Цех 1" } ] }
      ]
    }
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)

    user_id = int(user_data["id"])
    try:
        profile = UserProfile.objects.get(user_id=user_id)
    except UserProfile.DoesNotExist:
        return Response({"ok": True, "days": []})

    qs = (
        DayAssignment.objects
        .select_related("plan_day", "area")
        .filter(responsible=profile, plan_day__date__gte=date.today())
        .order_by("plan_day__date", "area__id")
    )

    grouped: dict[str, dict] = {}
    for item in qs:
        pd = item.plan_day
        date_str = pd.date.isoformat()
        if date_str not in grouped:
            grouped[date_str] = {
                "date": date_str,
                "passed": bool(pd.passed),
                "checked_at": pd.checked_at.isoformat() if pd.checked_at else None,
                "areas": [],
            }
        grouped[date_str]["areas"].append({
            "id": item.area.id,
            "name": item.area.name,
            "passed": bool(item.passed),
            "checked_at": item.checked_at.isoformat() if item.checked_at else None,
            "support_time": item.support_time.strftime("%H:%M") if item.support_time else None,
        })

    days = list(grouped.values())
    return Response({"ok": True, "days": days})

@api_view(["GET"])
def my_support_days(request):
    """
    Возвращает дни (и участки) для текущего пользователя как для Сопровождающего (watcher).
    Только с сегодняшней даты и дальше.
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    try:
        profile = UserProfile.objects.get(user_id=int(user_data["id"]))
    except UserProfile.DoesNotExist:
        return Response({"ok": True, "days": []})

    qs = (
        DayAssignment.objects
        .select_related("plan_day", "area")
        .filter(area__watcher=profile, plan_day__date__gte=date.today())
        .order_by("plan_day__date", "area__id")
    )

    grouped: dict[str, dict] = {}
    for item in qs:
        pd = item.plan_day
        date_str = pd.date.isoformat()
        if date_str not in grouped:
            grouped[date_str] = {
                "date": date_str,
                "passed": bool(pd.passed),
                "checked_at": pd.checked_at.isoformat() if pd.checked_at else None,
                "areas": [],
            }
        grouped[date_str]["areas"].append({
            "id": item.area.id,
            "name": item.area.name,
            "passed": bool(item.passed),
            "checked_at": item.checked_at.isoformat() if item.checked_at else None,
            "support_time": item.support_time.strftime("%H:%M") if item.support_time else None,
        })

    days = list(grouped.values())
    return Response({"ok": True, "days": days})

@api_view(["POST"])
def set_support_time(request):
    """
    Устанавливает время сопровождающего для конкретного участка в конкретный день.
    Body: { "date": "YYYY-MM-DD", "area_id": 1, "time": "HH:MM" }
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    body = request.data or {}
    date_str = (body.get("date") or "").strip()
    area_id = body.get("area_id")
    time_str = (body.get("time") or "").strip()
    if not date_str or not area_id or not time_str:
        return Response({"error": "date, area_id and time are required"}, status=status.HTTP_400_BAD_REQUEST)
    d = parse_date(date_str)
    t = parse_time(time_str)
    if not d or not t:
        return Response({"error": "invalid date or time"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        profile = UserProfile.objects.get(user_id=int(user_data["id"]))
    except UserProfile.DoesNotExist:
        return Response({"error": "profile not found"}, status=status.HTTP_404_NOT_FOUND)

    try:
        assignment = (
            DayAssignment.objects
            .select_related("plan_day", "area")
            .get(plan_day__date=d, area_id=area_id)
        )
    except DayAssignment.DoesNotExist:
        return Response({"error": "assignment not found"}, status=status.HTTP_404_NOT_FOUND)

    # Optionally restrict to assigned checker:
    if assignment.responsible_id and assignment.responsible_id != profile.id and profile.role != "admin":
        return Response({"error": "forbidden"}, status=status.HTTP_403_FORBIDDEN)

    assignment.support_time = t
    assignment.save(update_fields=["support_time"])
    return Response({"ok": True})


@api_view(["GET"])
def month_matrix(request):
    """
    Возвращает матрицу по месяцам: строки — участки, столбцы — дни месяца.
    Каждая ячейка: yes/total (кол-во ответов Да / всего вопросов чек-листа на момент обхода, если не было — текущее число вопросов).
    Query: ?year=YYYY&month=MM
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    try:
        y = int(request.query_params.get("year") or datetime.today().year)
        m = int(request.query_params.get("month") or datetime.today().month)
        first_day = date(y, m, 1)
        import calendar
        last_day = date(y, m, calendar.monthrange(y, m)[1])
    except Exception:
        return Response({"error": "invalid year/month"}, status=status.HTTP_400_BAD_REQUEST)

    # Lazy imports to avoid top circularities
    from .models import Checklist, ChecklistQuestion, InspectionSession, InspectionAnswer, PlanDay
    # Build days list
    days = []
    dcur = first_day
    while dcur <= last_day:
        days.append(dcur.isoformat())
        dcur = date.fromordinal(dcur.toordinal() + 1)

    # Areas
    areas = list(Area.objects.all().order_by("name").values("id", "name", "description"))

    # Pre-compute current total questions per area (fallback)
    area_total = {}
    for a in areas:
        checklist = Checklist.objects.filter(area_id=a["id"]).order_by("-id").first()
        total = ChecklistQuestion.objects.filter(checklist=checklist).count() if checklist else 0
        area_total[a["id"]] = total

    # Preload sessions of this month
    sessions = (InspectionSession.objects
                .select_related("area", "plan_day")
                .filter(plan_day__date__gte=first_day, plan_day__date__lte=last_day))
    # Map (area_id, date) -> session_ids
    from collections import defaultdict
    area_date_to_sessions: dict[tuple[int, str], list[int]] = defaultdict(list)
    for s in sessions:
        area_date_to_sessions[(s.area_id, s.plan_day.date.isoformat())].append(s.id)

    # Preload answers for these sessions
    answers = InspectionAnswer.objects.filter(session_id__in=[sid for sids in area_date_to_sessions.values() for sid in sids])
    # Map session_id -> (yes_count, total)
    from collections import Counter
    yes_count_by_session = Counter()
    total_by_session = Counter()
    for ans in answers:
        total_by_session[ans.session_id] += 1
        if ans.passed:
            yes_count_by_session[ans.session_id] += 1

    # Build data
    data = {}
    for a in areas:
        aid = a["id"]
        row = {}
        for di in days:
            sids = area_date_to_sessions.get((aid, di), [])
            if sids:
                yes_sum = sum(yes_count_by_session.get(sid, 0) for sid in sids)
                total_sum = sum(total_by_session.get(sid, 0) for sid in sids)
                # If multiple sessions in a day, aggregate
                row[di] = {"yes": yes_sum, "total": total_sum, "inspected": True}
            else:
                row[di] = {"yes": 0, "total": area_total.get(aid, 0), "inspected": False}
        data[str(aid)] = row

    return Response({
        "ok": True,
        "year": y,
        "month": m,
        "days": days,
        "areas": areas,
        "data": data,
    })


@api_view(["GET"])
def available_months(request):
    """
    Возвращает список месяцев, в которых есть хотя бы одна запись (DayAssignment).
    Формат: { ok, months: [ { year, month, count } ] } — отсортировано по убыванию.
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)

    agg = (DayAssignment.objects
           .values(year=ExtractYear("plan_day__date"), month=ExtractMonth("plan_day__date"))
           .annotate(count=Count("id"))
           .order_by("-year", "-month"))
    months = [{"year": a["year"], "month": a["month"], "count": a["count"]} for a in agg if a["year"] and a["month"]]
    return Response({"ok": True, "months": months})


@api_view(["POST"])
def month_matrix_export(request):
    """
    Генерирует excel-файл по отчёту за месяц и отправляет его в чат пользователю (Telegram).
    Body/Query: year, month
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    try:
        y = int(request.data.get("year") or request.query_params.get("year") or datetime.today().year)
        m = int(request.data.get("month") or request.query_params.get("month") or datetime.today().month)
    except Exception:
        return Response({"error": "invalid year/month"}, status=status.HTTP_400_BAD_REQUEST)

    # Build report data (reuse logic from month_matrix)
    from .models import Checklist, ChecklistQuestion, InspectionSession, InspectionAnswer, PlanDay
    import calendar
    first_day = date(y, m, 1)
    last_day = date(y, m, calendar.monthrange(y, m)[1])
    days = []
    dcur = first_day
    while dcur <= last_day:
        days.append(dcur.isoformat())
        dcur = date.fromordinal(dcur.toordinal() + 1)
    # Use only workdays (Mon-Fri) for the export
    work_days = [di for di in days if date.fromisoformat(di).weekday() < 5]

    areas = list(Area.objects.all().order_by("name").values("id", "name", "description"))
    # Pre-compute question totals by area (fallback)
    area_total = {}
    for a in areas:
        checklist = Checklist.objects.filter(area_id=a["id"]).order_by("-id").first()
        total = ChecklistQuestion.objects.filter(checklist=checklist).count() if checklist else 0
        area_total[a["id"]] = total

    # Sessions and answers map
    sessions = (InspectionSession.objects.select_related("area", "plan_day")
                .filter(plan_day__date__gte=first_day, plan_day__date__lte=last_day))
    from collections import defaultdict, Counter
    area_date_to_sessions: dict[tuple[int, str], list[int]] = defaultdict(list)
    for s in sessions:
        area_date_to_sessions[(s.area_id, s.plan_day.date.isoformat())].append(s.id)
    answers = InspectionAnswer.objects.filter(session_id__in=[sid for sids in area_date_to_sessions.values() for sid in sids])
    yes_count_by_session = Counter()
    total_by_session = Counter()
    for ans in answers:
        total_by_session[ans.session_id] += 1
        if ans.passed:
            yes_count_by_session[ans.session_id] += 1

    data = {}
    for a in areas:
        aid = a["id"]
        row = {}
        for di in days:
            sids = area_date_to_sessions.get((aid, di), [])
            if sids:
                yes_sum = sum(yes_count_by_session.get(sid, 0) for sid in sids)
                total_sum = sum(total_by_session.get(sid, 0) for sid in sids)
                row[di] = {"yes": yes_sum, "total": total_sum, "inspected": True}
            else:
                row[di] = {"yes": 0, "total": area_total.get(aid, 0), "inspected": False}
        data[str(aid)] = row

    # Build Excel
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Helper styles
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="1D4ED8")  # blue-700
    subtitle_fill = PatternFill("solid", fgColor="BFDBFE")  # blue-200
    red_fill = PatternFill("solid", fgColor="FDE2E2")

    # Widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36
    # dynamic day columns D .. D+len(work_days)-1
    for i in range(len(work_days)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 8
    # last 3 columns wider (x2)
    ws.column_dimensions[get_column_letter(4 + len(work_days))].width = 22
    ws.column_dimensions[get_column_letter(5 + len(work_days))].width = 22
    ws.column_dimensions[get_column_letter(6 + len(work_days))].width = 22

    # Title row
    total_cols = 3 + len(work_days) + 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value="Аналитика по культуре производства за месяц")
    c.font = Font(color="FFFFFF", bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = title_fill

    # Header group row
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
    ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=3 + len(work_days))
    ws.merge_cells(start_row=2, start_column=4 + len(work_days), end_row=3, end_column=6 + len(work_days))

    ws.cell(row=2, column=1, value="№ п/п").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=2, value="Участок").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=3, value="Пояснения").alignment = Alignment(horizontal="center", vertical="center")

    h = ws.cell(row=2, column=4, value="ОЦЕНКА ЗА ДЕНЬ")
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.fill = subtitle_fill

    h2 = ws.cell(row=2, column=4 + len(work_days), value="ИТОГОВАЯ ОЦЕНКА ЗА МЕСЯЦ (ДИНАМИКА)")
    h2.alignment = Alignment(horizontal="center", vertical="center")
    h2.fill = subtitle_fill

    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=3 + len(work_days))
    ws.cell(row=3, column=4, value="Число месяца").alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=4, column=4 + len(work_days), value="за текущий месяц").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=5 + len(work_days), value="за предыдущий месяц").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=6 + len(work_days), value="динамика изменения").alignment = Alignment(horizontal="center", vertical="center")

    # Day numbers row
    for i, di in enumerate(work_days):
        col = 4 + i
        ws.cell(row=4, column=col, value=int(di[-2:])).alignment = Alignment(horizontal="center", vertical="center")

    # Prepare previous month aggregates (Mon-Fri only)
    if m == 1:
        prev_y, prev_m = y - 1, 12
    else:
        prev_y, prev_m = y, m - 1
    prev_first_day = date(prev_y, prev_m, 1)
    prev_last_day = date(prev_y, prev_m, calendar.monthrange(prev_y, prev_m)[1])
    prev_days = []
    dcurp = prev_first_day
    while dcurp <= prev_last_day:
        prev_days.append(dcurp.isoformat())
        dcurp = date.fromordinal(dcurp.toordinal() + 1)
    prev_work_days = [di for di in prev_days if date.fromisoformat(di).weekday() < 5]

    sessions_prev = (InspectionSession.objects.select_related("area", "plan_day")
                     .filter(plan_day__date__gte=prev_first_day, plan_day__date__lte=prev_last_day))
    from collections import defaultdict as _defaultdict, Counter as _Counter
    area_date_to_sessions_prev: dict[tuple[int, str], list[int]] = _defaultdict(list)
    for s in sessions_prev:
        area_date_to_sessions_prev[(s.area_id, s.plan_day.date.isoformat())].append(s.id)
    answers_prev = InspectionAnswer.objects.filter(session_id__in=[sid for sids in area_date_to_sessions_prev.values() for sid in sids])
    yes_count_by_session_prev = _Counter()
    total_by_session_prev = _Counter()
    for ans in answers_prev:
        total_by_session_prev[ans.session_id] += 1
        if ans.passed:
            yes_count_by_session_prev[ans.session_id] += 1

    # Body
    row_idx = 5
    for idx, a in enumerate(areas, start=1):
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=a["name"])
        ws.cell(row=row_idx, column=3, value=a.get("description") or "")
        # make row taller
        ws.row_dimensions[row_idx].height = 26
        # day scores and avg
        curr_sum = 0.0
        curr_cnt = 0
        for i, di in enumerate(work_days):
            col = 4 + i
            cdata = data[str(a["id"])][di]
            inspected = cdata["inspected"]
            yes = cdata["yes"]
            total = cdata["total"]
            score = int(round((yes / total) * 10)) if inspected and total > 0 else 0
            cell = ws.cell(row=row_idx, column=col, value=score)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # colors
            if not inspected:
                cell.fill = red_fill
            else:
                if score >= 8:
                    cell.font = Font(color="006400", bold=True)  # green
                elif score >= 6:
                    cell.font = Font(color="9A7B00", bold=True)  # yellow-ish
                else:
                    cell.font = Font(color="8B0000", bold=True)  # red
                curr_sum += score
                curr_cnt += 1
        curr_avg = round(curr_sum / curr_cnt, 1) if curr_cnt else 0.0
        # previous month average (Mon-Fri inspected days only)
        prev_avg = 0.0
        p_sum = 0.0
        p_cnt = 0
        for di in prev_work_days:
            sids = area_date_to_sessions_prev.get((a["id"], di), [])
            if sids:
                yes_sum = sum(yes_count_by_session_prev.get(sid, 0) for sid in sids)
                total_sum = sum(total_by_session_prev.get(sid, 0) for sid in sids)
                if total_sum > 0:
                    p_sum += (yes_sum / total_sum) * 10
                    p_cnt += 1
        prev_avg = round(p_sum / p_cnt, 1) if p_cnt else 0.0

        ws.cell(row=row_idx, column=4 + len(work_days), value=curr_avg).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=5 + len(work_days), value=prev_avg or None).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=6 + len(work_days), value=(round(curr_avg - prev_avg, 1) if p_cnt else None)).alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1

    # Borders for all used cells
    for r in ws.iter_rows(min_row=1, max_row=row_idx - 1, min_col=1, max_col=total_cols):
        for cell in r:
            cell.border = border

    # Save to bytes
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"Культура_Производства_Месяц_{y}-{str(m).zfill(2)}.xlsx"

    # Send to chat
    try:
        chat_id = int(user_data["id"])
        sent = send_excel_report(chat_id=chat_id, filename=fname, data=bio.getvalue(), caption=fname)
        return Response({"ok": True, "sent": bool(sent)})
    except Exception:
        return Response({"ok": False, "sent": False})
@api_view(["GET"])
def checker_day_detail(request):
    """
    Возвращает подробности по обходу на конкретную дату и участок.
    Query: ?date=YYYY-MM-DD&area_id=1
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    date_str = (request.query_params.get("date") or "").strip()
    area_id = request.query_params.get("area_id")
    if not date_str or not area_id:
        return Response({"error": "date and area_id are required"}, status=status.HTTP_400_BAD_REQUEST)
    d = parse_date(date_str)
    if not d:
        return Response({"error": "invalid date"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        assignment = (
            DayAssignment.objects
            .select_related("plan_day", "area", "responsible", "area__watcher")
            .get(plan_day__date=d, area_id=area_id)
        )
    except DayAssignment.DoesNotExist:
        return Response({"error": "not found"}, status=status.HTTP_404_NOT_FOUND)

    area = assignment.area
    watcher = area.watcher
    responsible = assignment.responsible

    data = {
        "date": assignment.plan_day.date.isoformat(),
        "day": {
            "passed": bool(assignment.plan_day.passed),
            "checked_at": assignment.plan_day.checked_at.isoformat() if assignment.plan_day.checked_at else None,
        },
        "area": {
            "id": area.id,
            "name": area.name,
            "watcher": {
                "user_id": watcher.user_id if watcher else None,
                "full_name": watcher.full_name if watcher else None,
                "role": watcher.role if watcher else None,
            } if watcher else None,
        },
        "assignment": {
            "passed": bool(assignment.passed),
            "checked_at": assignment.checked_at.isoformat() if assignment.checked_at else None,
            "support_time": assignment.support_time.strftime("%H:%M") if assignment.support_time else None,
            "responsible": {
                "user_id": responsible.user_id if responsible else None,
                "full_name": responsible.full_name if responsible else None,
                "role": responsible.role if responsible else None,
            } if responsible else None,
        },
        "proposals": [
            {
                "id": p.id,
                "proposed_time": p.proposed_time.strftime("%H:%M"),
                "status": p.status,
                "proposed_by": {
                    "user_id": p.proposed_by.user_id if p.proposed_by else None,
                    "full_name": p.proposed_by.full_name if p.proposed_by else None,
                },
                "decided_at": p.decided_at.isoformat() if p.decided_at else None,
            }
            for p in SupportProposal.objects.filter(plan_day__date=d, area_id=area_id).order_by("-created_at")[:10]
        ],
        "start_allowed": (assignment.checked_at is None and not assignment.passed and assignment.support_time is not None),
    }
    # attach last inspection summary if exists
    from .models import InspectionSession, InspectionAnswer, ChecklistQuestion
    last_session = (InspectionSession.objects
                    .filter(plan_day=assignment.plan_day, area=assignment.area, completed_at__isnull=False)
                    .order_by("-completed_at", "-started_at")
                    .first())
    if last_session:
        answers = (InspectionAnswer.objects
                   .select_related("question")
                   .filter(session=last_session)
                   .order_by("answered_at"))
        # Build answers with public media host
        last_answers = []
        for ans in answers:
            photo_url = None
            if ans.defect_photo:
                raw = ans.defect_photo.url
                photo_url = make_public_url(raw)
            last_answers.append({
                "question_id": ans.question_id,
                "question_text": ans.question.text,
                "passed": bool(ans.passed),
                "defect_photo": photo_url,
            })
        data["last_inspection"] = {
            "completed_at": last_session.completed_at.isoformat() if last_session.completed_at else None,
            "answers": last_answers
        }
    return Response({"ok": True, "detail": data})


@api_view(["POST"])
def propose_support_time(request):
    """
    Checker proposes time to watcher for a specific day+area.
    Body: { "date": "YYYY-MM-DD", "area_id": 1, "time": "HH:MM" }
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    body = request.data or {}
    date_str = (body.get("date") or "").strip()
    area_id = body.get("area_id")
    time_str = (body.get("time") or "").strip()
    if not date_str or not area_id or not time_str:
        return Response({"error": "date, area_id and time are required"}, status=status.HTTP_400_BAD_REQUEST)
    d = parse_date(date_str)
    t = parse_time(time_str)
    if not d or not t:
        return Response({"error": "invalid date or time"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        profile = UserProfile.objects.get(user_id=int(user_data["id"]))
    except UserProfile.DoesNotExist:
        return Response({"error": "profile not found"}, status=status.HTTP_404_NOT_FOUND)

    # Ensure assignment exists
    try:
        _ = DayAssignment.objects.get(plan_day__date=d, area_id=area_id)
    except DayAssignment.DoesNotExist:
        return Response({"error": "assignment not found"}, status=status.HTTP_404_NOT_FOUND)

    sp = SupportProposal.objects.create(
        plan_day_id=DayAssignment.objects.filter(plan_day__date=d, area_id=area_id).values_list("plan_day_id", flat=True).first(),
        area_id=area_id,
        proposed_by=profile,
        proposed_time=t,
        status="pending",
    )
    # Notify watcher
    try:
        assignment = DayAssignment.objects.select_related("area", "area__watcher").get(plan_day__date=d, area_id=area_id)
        watcher = assignment.area.watcher
        if watcher and watcher.user_id:
            from .services.telegram_notifier import notify_support_proposal
            proposer_name = profile.full_name or str(profile.user_id)
            notify_support_proposal(
                watcher_user_id=int(watcher.user_id),
                area_name=assignment.area.name,
                date_iso=d.isoformat(),
                time_str=t.strftime("%H:%M"),
                proposer_full_name=proposer_name,
            )
    except Exception:
        pass
    return Response({"ok": True, "proposal_id": sp.id})


@api_view(["POST"])
def respond_support_proposal(request):
    """
    Watcher/admin responds to proposal.
    Body: { "proposal_id": N, "action": "accept"|"reject" }
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    body = request.data or {}
    pid = body.get("proposal_id")
    action = (body.get("action") or "").strip().lower()
    if not pid or action not in ("accept", "reject"):
        return Response({"error": "proposal_id and valid action are required"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        responder = UserProfile.objects.get(user_id=int(user_data["id"]))
    except UserProfile.DoesNotExist:
        return Response({"error": "profile not found"}, status=status.HTTP_404_NOT_FOUND)
    try:
        sp = SupportProposal.objects.select_related("area", "plan_day", "area__watcher").get(id=pid)
    except SupportProposal.DoesNotExist:
        return Response({"error": "proposal not found"}, status=status.HTTP_404_NOT_FOUND)

    # Only watcher for area or admin can respond
    if responder.role != "admin":
        if not sp.area.watcher or sp.area.watcher_id != responder.id:
            return Response({"error": "forbidden"}, status=status.HTTP_403_FORBIDDEN)

    sp.status = "accepted" if action == "accept" else "rejected"
    sp.decided_by = responder
    sp.decided_at = timezone.now()
    sp.save(update_fields=["status", "decided_by", "decided_at"])

    # On accept, set assignment.support_time
    if action == "accept":
        try:
            assignment = DayAssignment.objects.get(plan_day=sp.plan_day, area=sp.area)
            assignment.support_time = sp.proposed_time
            assignment.save(update_fields=["support_time"])
            # Auto-reject other pending proposals for the same day and area
            (SupportProposal.objects
             .filter(plan_day=sp.plan_day, area=sp.area, status="pending")
             .exclude(id=sp.id)
             .update(status="rejected", decided_by_id=responder.id, decided_at=timezone.now()))
        except DayAssignment.DoesNotExist:
            pass

    # Notify proposer about decision
    try:
        if sp.proposed_by and sp.proposed_by.user_id:
            from .services.telegram_notifier import notify_proposal_response
            notify_proposal_response(
                proposer_user_id=int(sp.proposed_by.user_id),
                area_name=sp.area.name,
                date_iso=sp.plan_day.date.isoformat(),
                time_str=sp.proposed_time.strftime("%H:%M") if sp.proposed_time else "",
                action=action,
                decided_by_full_name=responder.full_name if responder else None,
            )
    except Exception:
        pass

    return Response({"ok": True, "status": sp.status})


@api_view(["POST"])
def start_inspection(request):
    """
    Начинает сессию обхода для checker.
    Body: { "date": "YYYY-MM-DD", "area_id": 1 }
    Returns: { ok, session_id, questions: [{id, text, reference_image}] }
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    body = request.data or {}
    date_str = (body.get("date") or "").strip()
    area_id = body.get("area_id")
    if not date_str or not area_id:
        return Response({"error": "date and area_id are required"}, status=status.HTTP_400_BAD_REQUEST)
    d = parse_date(date_str)
    if not d:
        return Response({"error": "invalid date"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        profile = UserProfile.objects.get(user_id=int(user_data["id"]))
    except UserProfile.DoesNotExist:
        return Response({"error": "profile not found"}, status=status.HTTP_404_NOT_FOUND)

    try:
        assignment = DayAssignment.objects.select_related("plan_day", "area").get(plan_day__date=d, area_id=area_id)
    except DayAssignment.DoesNotExist:
        return Response({"error": "assignment not found"}, status=status.HTTP_404_NOT_FOUND)

    from .models import InspectionSession, ChecklistQuestion
    # Find checklist for area
    checklist = Checklist.objects.filter(area_id=area_id).order_by("-id").first() if 'Checklist' in globals() else None
    if not checklist:
        from .models import Checklist
        checklist = Checklist.objects.filter(area_id=area_id).order_by("-id").first()
    if not checklist:
        return Response({"error": "no checklist for area"}, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        session = InspectionSession.objects.create(plan_day=assignment.plan_day, area=assignment.area, checker=profile)
    # Previous issues from last completed session for this area before the current day
    from django.utils import timezone as _tz
    last_session = (InspectionSession.objects
                    .filter(area=assignment.area, completed_at__isnull=False, plan_day__date__lt=assignment.plan_day.date)
                    .order_by("-completed_at", "-started_at")
                    .first())
    prev_fail_map = {}
    if last_session:
        from .models import InspectionAnswer as _Ans, ChecklistQuestion as _Q
        prev_answers = (_Ans.objects
                        .select_related("question")
                        .filter(session=last_session, passed=False))
        for ans in prev_answers:
            photo_url = None
            if ans.defect_photo:
                raw = ans.defect_photo.url
                photo_url = make_public_url(raw)
            prev_fail_map[ans.question_id] = {
                "prev_failed": True,
                "prev_defect_photo": photo_url,
                "prev_answered_at": ans.answered_at.isoformat() if ans.answered_at else None,
            }
    # Build questions
    questions = []
    for q in ChecklistQuestion.objects.filter(checklist=checklist).order_by("order", "id"):
        ref_url = q.reference_image.url if q.reference_image else None
        if ref_url:
            ref_url = make_public_url(ref_url)
        qpayload = {"id": q.id, "text": q.text, "reference_image": ref_url}
        if q.id in prev_fail_map:
            qpayload.update(prev_fail_map[q.id])
        questions.append(qpayload)
    return Response({"ok": True, "session_id": session.id, "questions": questions})


@api_view(["POST"])
def answer_inspection(request):
    """
    Сохраняет ответ по вопросу обхода.
    Multipart/FormData: session_id, question_id, passed (true/false), defect_photo (optional)
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    try:
        from .models import InspectionSession, InspectionAnswer, ChecklistQuestion
        session_id = int(request.data.get("session_id"))
        question_id = int(request.data.get("question_id"))
        passed_val = str(request.data.get("passed", "true")).lower() in ("1", "true", "yes")
    except Exception:
        return Response({"error": "invalid payload"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        session = InspectionSession.objects.select_related("checker").get(id=session_id)
    except InspectionSession.DoesNotExist:
        return Response({"error": "session not found"}, status=status.HTTP_404_NOT_FOUND)
    # Optional: ensure same checker
    if session.checker and session.checker.user_id != int(user_data["id"]):
        return Response({"error": "forbidden"}, status=status.HTTP_403_FORBIDDEN)
    try:
        question = ChecklistQuestion.objects.get(id=question_id)
    except ChecklistQuestion.DoesNotExist:
        return Response({"error": "question not found"}, status=status.HTTP_404_NOT_FOUND)
    photo = request.FILES.get("defect_photo")
    InspectionAnswer.objects.create(session=session, question=question, passed=passed_val, defect_photo=photo)
    return Response({"ok": True})


@api_view(["POST"])
def complete_inspection(request):
    """
    Завершает сессию обхода, проставляет статусы.
    Body: { "session_id": N }
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    try:
        from .models import InspectionSession
        session_id = int(request.data.get("session_id"))
    except Exception:
        return Response({"error": "invalid payload"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        session = InspectionSession.objects.select_related("plan_day", "area").get(id=session_id)
    except InspectionSession.DoesNotExist:
        return Response({"error": "session not found"}, status=status.HTTP_404_NOT_FOUND)
    if session.checker and session.checker.user_id != int(user_data["id"]):
        return Response({"error": "forbidden"}, status=status.HTTP_403_FORBIDDEN)
    session.completed_at = timezone.now()
    session.save(update_fields=["completed_at"])
    # Mark DayAssignment passed if all answers passed
    try:
        da = DayAssignment.objects.get(plan_day=session.plan_day, area=session.area)
        # Считаем "пройдено" по факту завершения обхода, независимо от ответов "нет"
        da.passed = True
        da.checked_at = timezone.now()
        da.save(update_fields=["passed", "checked_at"])
        # Mark the day as passed when there is at least one completed assignment
        plan = session.plan_day
        if not plan.passed:
            if plan.assignments.filter(checked_at__isnull=False).exists():
                plan.passed = True
                plan.checked_at = timezone.now()
                plan.save(update_fields=["passed", "checked_at"])
    except DayAssignment.DoesNotExist:
        pass
    return Response({"ok": True})


@api_view(["GET"])
def analytics_overview(request):
    """
    Агрегированная аналитика:
    - heatmap: оценки 0-10 по (участок x дата)
    - top_failures: топ часто проваливаемых вопросов
    - trends: тренд нарушений по датам (fail_rate, fails, total, avg_score)
    - schedule: выполнение графика по неделям (scheduled vs inspected, %)
    Параметры: ?from=YYYY-MM-DD&to=YYYY-MM-DD (по умолчанию последние 30 дней), только рабочие дни (Пн–Пт).
    """
    user_data = getattr(request, "user_data", None)
    if not user_data or not user_data.get("id"):
        return Response({"error": "Unauthorized"}, status=status.HTTP_401_UNAUTHORIZED)
    from datetime import timedelta
    from collections import defaultdict, Counter
    from .models import InspectionSession, InspectionAnswer, ChecklistQuestion

    # Дата диапазон
    to_str = (request.query_params.get("to") or "").strip()
    from_str = (request.query_params.get("from") or "").strip()
    to_date = parse_date(to_str) if to_str else date.today()
    if not to_date:
        to_date = date.today()
    from_date = parse_date(from_str) if from_str else (to_date - timedelta(days=29))
    if not from_date:
        from_date = to_date - timedelta(days=29)
    if from_date > to_date:
        from_date, to_date = to_date, from_date

    # Рабочие дни
    days = []
    dcur = from_date
    while dcur <= to_date:
        if dcur.weekday() < 5:
            days.append(dcur)
        dcur = dcur + timedelta(days=1)
    days_iso = [d.isoformat() for d in days]

    # Справочники
    areas = list(Area.objects.all().order_by("name").values("id", "name", "description"))

    # Сессии и ответы
    sessions = (InspectionSession.objects.select_related("area", "plan_day")
                .filter(plan_day__date__gte=from_date, plan_day__date__lte=to_date))
    area_date_to_sessions: dict[tuple[int, str], list[int]] = defaultdict(list)
    for s in sessions:
        area_date_to_sessions[(s.area_id, s.plan_day.date.isoformat())].append(s.id)
    all_session_ids = [sid for sids in area_date_to_sessions.values() for sid in sids]
    answers = (InspectionAnswer.objects
               .filter(session_id__in=all_session_ids)
               .select_related("question", "session", "session__plan_day", "session__area"))

    yes_by_session = Counter()
    total_by_session = Counter()
    for ans in answers:
        total_by_session[ans.session_id] += 1
        if ans.passed:
            yes_by_session[ans.session_id] += 1

    # Heatmap
    heatmap = {}
    for a in areas:
        aid = a["id"]
        row = {}
        for di in days_iso:
            sids = area_date_to_sessions.get((aid, di), [])
            if sids:
                yes_sum = sum(yes_by_session.get(sid, 0) for sid in sids)
                total_sum = sum(total_by_session.get(sid, 0) for sid in sids)
                score = int(round((yes_sum / total_sum) * 10)) if total_sum > 0 else 0
                row[di] = {"score": score, "inspected": True}
            else:
                row[di] = {"score": 0, "inspected": False}
        heatmap[str(aid)] = row

    # Top failures
    top_qs = (InspectionAnswer.objects
              .filter(session_id__in=all_session_ids, passed=False)
              .values("question_id", "question__text", "session__area__name")
              .annotate(count=Count("id"), last_failed_at=Max("answered_at"))
              .order_by("-count")[:6])
    top_failures = [{
        "question_id": t["question_id"],
        "question_text": t["question__text"],
        "area_name": t["session__area__name"],
        "count": t["count"],
        "last_failed_at": (t["last_failed_at"].isoformat() if t["last_failed_at"] else None)
    } for t in top_qs]

    # Trends
    date_fails = Counter()
    date_total = Counter()
    for ans in answers:
        di = ans.session.plan_day.date.isoformat()
        if di not in days_iso:
            continue
        date_total[di] += 1
        if not ans.passed:
            date_fails[di] += 1
    trends = []
    for di in days_iso:
        tot = date_total.get(di, 0)
        fails = date_fails.get(di, 0)
        rate = (fails / tot) if tot > 0 else 0.0
        scores = []
        for a in areas:
            cell = heatmap[str(a["id"])][di]
            if cell["inspected"]:
                scores.append(cell["score"])
        avg_score = round(sum(scores) / len(scores), 1) if scores else 0.0
        trends.append({
            "date": di,
            "fails": fails,
            "total": tot,
            "fail_rate": round(rate, 3),
            "avg_score": avg_score
        })

    # Schedule adherence
    assignments = DayAssignment.objects.select_related("plan_day", "area").filter(
        plan_day__date__gte=from_date, plan_day__date__lte=to_date
    )
    week_map = defaultdict(lambda: {"scheduled": 0, "inspected": 0})
    for aobj in assignments:
        d = aobj.plan_day.date
        week_start = (d - timedelta(days=d.weekday()))  # Monday
        wk = week_start.isoformat()
        week_map[wk]["scheduled"] += 1
        if area_date_to_sessions.get((aobj.area_id, d.isoformat())):
            week_map[wk]["inspected"] += 1
    schedule = []
    for wk in sorted(week_map.keys()):
        sch = week_map[wk]["scheduled"]
        ins = week_map[wk]["inspected"]
        adherence = round((ins / sch) * 100, 1) if sch else 0.0
        schedule.append({
            "week_start": wk,
            "scheduled": sch,
            "inspected": ins,
            "adherence": adherence
        })

    return Response({
        "ok": True,
        "range": {"from": from_date.isoformat(), "to": to_date.isoformat()},
        "days": days_iso,
        "areas": [{"id": a["id"], "name": a["name"]} for a in areas],
        "heatmap": heatmap,
        "top_failures": top_failures,
        "trends": trends,
        "schedule": schedule
    })
